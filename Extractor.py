import os
import re
import json
import base64
import pandas as pd
import time
import random
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional

from ibm_watsonx_ai import Credentials
from ibm_watsonx_ai.foundation_models import ModelInference
from image_processor import enhance_image_for_ocr, process_images_in_folder
from ibm_watsonx_ai import APIClient
from concurrent.futures import ThreadPoolExecutor, as_completed

##########################################
### LOGGING SETUP
##########################################

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('extraction_log.txt'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

##########################################
### PATHS & CONFIGURATION
##########################################

RAW_IMAGES_DIR = r"C:\\Users\\LENOVO\Desktop\\Invoice Extraction\\INPUT_IMAGES\\Catnip New PoC Samples\\Invoice_18-Nov-2025"
PREPROCESSED_DIR = r"C:\\Users\\LENOVO\Desktop\\Invoice Extraction\\PROCESSED_IMAGES"
EXCEL_OUTPUT_PATH = r"C:\\Users\\LENOVO\Desktop\\Invoice Extraction\\extracted_data.xlsx"
IMAGES_DIR = PREPROCESSED_DIR

API_KEY = "9dhWazh67-KWTUjf80FnPsmEUx828u1vid3PNrFAmuSS"
SERVICE_URL = "https://us-south.ml.cloud.ibm.com/"
PROJECT_ID = "52cf4e44-2a8c-474d-b247-11957cc6891d"
MODEL_ID = "meta-llama/llama-4-maverick-17b-128e-instruct-fp8"

##########################################
### PREPROCESS IMAGES
##########################################

# print("🔧 Running Enhanced Image Preprocessor...")
# process_images_in_folder(RAW_IMAGES_DIR, PREPROCESSED_DIR, scale_factor=1.5)
# print("✅ Preprocessing Completed!\n")

##########################################
### WATSONX MODEL SETUP
##########################################

creds = Credentials(url=SERVICE_URL, api_key=API_KEY)
api_client = APIClient(creds)
api_client.set.default_project(PROJECT_ID)
model = ModelInference(api_client=api_client, model_id=MODEL_ID)

info = model.get_details()
print(f"✅ Connected to model: {info['model_id']}")

generation_params = {
    "max_new_tokens": 12000,  # Increased for complex invoices
    "temperature": 0.05,  # Very low for maximum consistency
    "top_p": 0.95,
    "repetition_penalty": 1.15  # Stronger penalty for repetition
}

##########################################
### VALIDATION PATTERNS
##########################################

# GST Number validation pattern
GST_PATTERN = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z][A-Z][0-9A-Z]$')

# Valid GST state codes
VALID_STATE_CODES = {
    '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
    '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
    '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
    '31', '32', '33', '34', '35', '36', '37'
}

# Valid tax percentages
VALID_TAX_PERCENTAGES = [0, 0.25, 1, 1.5, 3, 5, 6, 9, 12, 14, 18, 28]

# Phone number pattern (Indian)
PHONE_PATTERN = re.compile(r'(\+91[\-\s]?)?[6-9]\d{9}')

# IMEI pattern (15 digits)
IMEI_PATTERN = re.compile(r'\b\d{15}\b')

# Date patterns
DATE_PATTERNS = [
    r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # DD/MM/YYYY or DD-MM-YYYY
    r'\d{2,4}[/-]\d{1,2}[/-]\d{1,2}',  # YYYY/MM/DD
]

##########################################
### ENHANCED EXTRACTION PROMPT
##########################################

invoice_prompt = """
You are an expert OCR-based invoice data extractor with 99% accuracy requirement.

Extract ALL visible information from this invoice image with MAXIMUM PRECISION.

═══════════════════════════════════════════════════════════════
🔴 CRITICAL EXTRACTION RULES
═══════════════════════════════════════════════════════════════

1. **ACCURACY FIRST**: If unsure about any field, return "" rather than guessing
2. **VERIFY TWICE**: Double-check each extracted value before finalizing
3. **LOCATION AWARENESS**: Know where to look for each field type
4. **NO ASSUMPTIONS**: Only extract what is CLEARLY VISIBLE

═══════════════════════════════════════════════════════════════
📋 INVOICE NUMBER & DATE (CRITICAL - 100% Required)
═══════════════════════════════════════════════════════════════

**invoiceNumber**:
- Location: Top section, usually labeled "Invoice No", "Bill No", "Inv#"
- Can be: Numeric, alphanumeric, or with special characters
- Common formats: "INV-001", "2024/123", "BILL001"
- If handwritten, extract carefully character by character
- Return "" ONLY if absolutely not visible

**invoiceNumberType**:
- "Printed" if in standard typed/printed font
- "Handwritten" if written by hand
- Look at font consistency and style

**invoiceDate**:
- Location: Near invoice number, labeled "Date", "Invoice Date", "Bill Date"
- Extract EXACTLY as shown (DD/MM/YYYY, DD-MM-YYYY, etc.)
- Do NOT reformat or change the date format
- Return "" if not visible

═══════════════════════════════════════════════════════════════
🏢 SUPPLIER INFORMATION (CRITICAL)
═══════════════════════════════════════════════════════════════

**supplierName**:
- Location: TOP of invoice (header section)
- Usually the LARGEST text or company logo text
- May be in bold or different font
- Extract the full business name
- Do NOT include address or contact details here

**gstNumber** (GSTIN):
- MUST be EXACTLY 15 characters
- Should compulsarily match this Format: [0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z][A-Z][0-9A-Z]
- Example: 27AABCT1234F1Z5
- First 2 digits = State code (01-37)
- Validate format before returning
- Return "" if doesn't match pattern EXACTLY

═══════════════════════════════════════════════════════════════
👤 CUSTOMER INFORMATION
═══════════════════════════════════════════════════════════════

**customerName**:
- Location: Look for "Bill To", "Customer", "Buyer", "Sold To" section
- Usually in middle-left of invoice
- Extract full name (person or company)
- Return "" if not present

**customerPhone**:
- Look for phone/mobile number in customer section
- Format: Extract as-is (may include +91, spaces, hyphens)
- Must be 10 digits (Indian) or with country code
- Do NOT extract supplier's phone
- Return "" if not present

**customerAddress**:
- Complete address under customer/buyer section
- Include all lines (street, city, state, PIN)
- Return "" if not present

═══════════════════════════════════════════════════════════════
💰 FINANCIAL FIELDS
═══════════════════════════════════════════════════════════════

**downPayment**:
- Look for: "Down Payment", "DP", "Advance", "Deposit"
- Extract numeric value only (no currency symbols)
- Return "" if not present
- This value is usually present near the "Hypothecation" stamp or word 

**netTotal**:
- Look for: "Grand Total", "Net Total", "Total Amount", "Amount Payable"
- Usually at BOTTOM of invoice in summary section
- Extract the FINAL total amount
- Return "" if not clearly visible

═══════════════════════════════════════════════════════════════
📦 ITEM TABLE EXTRACTION (CRITICAL FOR MULTI-LINE INVOICES)
═══════════════════════════════════════════════════════════════

🔴 COLUMN IDENTIFICATION (EXTREMELY IMPORTANT):
In the items table, columns typically appear in this order from LEFT to RIGHT:

Position 1-2: S.No / Item No / Description
Position 3-4: HSN/SAC Code / Brand / Quantity
Position 5: **RATE** (Unit Price) ← This is the price per single unit
Position 6-8: Tax Percentage Columns (SGST%, CGST%, IGST%)
Position 9-10: **AMOUNT** (Final Line Total) ← This is Rate × Qty ± Tax

🔴 CRITICAL RULES FOR RATE vs AMOUNT:
1. **RATE** is ALWAYS the unit price (price for 1 item)
2. **AMOUNT** is ALWAYS the final calculated value for that line item
3. For multi-line invoices, Amount is usually the RIGHTMOST numeric column
4. Rate column appears BEFORE tax percentage columns
5. Amount column appears AFTER tax percentage columns
6. If Quantity > 1, then Amount should be LARGER than Rate
7. Amount = Rate × Quantity × (1 + Tax%)

🔴 VISUAL CUES TO IDENTIFY COLUMNS:
- Rate column header: "Rate", "Unit Price", "Price", "Basic Rate", "MRP"
- Amount column header: "Amount", "Total", "Line Total", "Value", "Taxable Value"
- Rate values are typically SMALLER numbers
- Amount values are typically LARGER numbers (especially when qty > 1)
- Amount column is usually RIGHT-ALIGNED in the table

Extract EVERY row from the items table. For each item:

**itemNo**: Sequential number (1, 2, 3...) or as shown

**description**:
- Full item/product description
- Include model numbers, specifications
- Do NOT include IMEI/Serial numbers here

**brandName** (MANDATORY):
- Extract brand from description or separate field
- Common: Samsung, Apple, Vivo, Oppo, Realme, Xiaomi, OnePlus, TVS, Honda, Hero, Bajaj
- This is CRITICAL - try your best to identify
- Return "" only if absolutely no brand visible

**imeiNumber**:
- 15-digit number for mobile phones
- Look for "IMEI", "IMEI1", "IMEI2" labels
- Can be in description or separate field
- Return "" if not present

**serialNumber**:
- Look for "Serial No", "S.No", "Chassis No", "Engine No"
- Alphanumeric code
- Return "" if not present

**quantity**:
- Number of units
- Usually labeled "Qty", "Quantity", "Nos"
- Extract numeric value
- Return "" if not visible

**rate**:
- Unit price per item (NOT the final amount)
- Column labeled: "Rate", "Unit Price", "Price", "Basic Rate"
- Extract from ITEMS TABLE ONLY (not from summary)
- This is the price for ONE unit, not the total
- Look for the column that comes BEFORE tax columns
- Do NOT confuse with: Amount, Tax%, Quantity, HSN code
- For multi-line invoices, scan carefully from left to identify the rate column
- Return "" if not clearly visible in rate column

**itemAmount**:
- Final amount for this line item (Rate × Qty with or without tax)
- Column labeled: "Amount", "Total", "Line Total", "Value"
- This is the RIGHTMOST numeric column in most cases
- This value should be LARGER than rate when quantity > 1
- Extract the exact value shown in the Amount column
- Do NOT calculate - extract what is printed
- For multi-line invoices, this is CRITICAL - look for the rightmost column
- Return "" if not visible

═══════════════════════════════════════════════════════════════
💸 TAX EXTRACTION (CRITICAL - Percentage ONLY)
═══════════════════════════════════════════════════════════════

**VALID TAX %**: 0, 0.25, 1, 1.5, 3, 5, 6, 9, 12, 14, 18, 28, 1.46,

**RULES**:
1. Extract PERCENTAGE ONLY (not amounts)
2. Must see % symbol next to number
3. Value must be in valid list above
4. If not in list → return ""

**EXAMPLES**:

✅ "GST @ 18%" → sgst: 9, cgst: 9, igst: ""
✅ "IGST 12%" → sgst: "", cgst: "", igst: 12
✅ "SGST 9%, CGST 9%" → sgst: 9, cgst: 9, igst: ""
❌ "Tax: 2578.50" → sgst: "", cgst: "", igst: "" (this is amount, not %)

**sgst**: State GST percentage
**cgst**: Central GST percentage
**igst**: Integrated GST percentage

**IMPORTANT**:
- If SGST/CGST present → igst must be ""
- If IGST present → sgst and cgst must be ""
- example: If only "GST 18%" visible → sgst: (visible GST% /2), cgst: (visible GST% /2), igst: ""

═══════════════════════════════════════════════════════════════
🔖 STAMP & SIGNATURE DETECTION
═══════════════════════════════════════════════════════════════

**CRITICAL**: Scan BOTTOM 30% of invoice thoroughly!

**stampPresent**:
- Look for company stamp/seal (circular or rectangular)
- Common location: Bottom-right corner
- May be in RED, BLUE, or BLACK ink
- Can be faded or light
- Return "Present" or "Absent"

**informationInStamp**:
- Extract ALL text inside the company stamp
- May include company name, GSTIN, address
- Text may be curved or at angles
- Return "" if no stamp

**signaturePresent**:
- Look for handwritten signature
- Usually near stamp or at bottom
- Return "Yes" or "No"

**hypothecationStamp**:
- Look for "Hypothecated to..." text
- This is DIFFERENT from company stamp
- Return "Present" or "Absent"

**stampCompanyMatching_score**:
- Compare supplierName with informationInStamp
- Return similarity score 0-100
- Return 0 if no stamp

═══════════════════════════════════════════════════════════════
📤 OUTPUT FORMAT
═══════════════════════════════════════════════════════════════

Return ONLY valid JSON. No markdown. No explanations. No extra text.

{
  "invoiceNumber": "",
  "invoiceNumberType": "",
  "invoiceDate": "",
  "supplierName": "",
  "gstNumber": "",
  "customerName": "",
  "customerPhone": "",
  "customerAddress": "",
  "downPayment": "",
  "netTotal": "",
  "stampPresent": "",
  "informationInStamp": "",
  "signaturePresent": "",
  "hypothecationStamp": "",
  "stampCompanyMatching_score": 0,
  "items": [
    {
      "itemNo": "",
      "description": "",
      "brandName": "",
      "imeiNumber": "",
      "serialNumber": "",
      "quantity": "",
      "rate": "",
      "sgst": "",
      "cgst": "",
      "igst": "",
      "tax": "",
      "itemAmount": ""
    }
  ]
}

═══════════════════════════════════════════════════════════════
✅ FINAL CHECKLIST
═══════════════════════════════════════════════════════════════

Before returning JSON, verify:
- [ ] Invoice number extracted correctly
- [ ] All items from table included
- [ ] Tax values are PERCENTAGES (not amounts)
- [ ] GST number is 15 characters (if present)
- [ ] Stamp checked in bottom section
- [ ] For multi-line invoices: Rate and Amount columns identified correctly
- [ ] Rate is unit price, Amount is line total
- [ ] All "" for missing values (no null, no "N/A")
"""

##########################################
### FIELD VALIDATORS
##########################################

def validate_gst_number(gst: str) -> str:
    """Validate GST number format and return cleaned value or empty string."""
    if not gst:
        return ""
    
    # Remove spaces and convert to uppercase
    gst = gst.strip().upper().replace(" ", "")
    
    # Check length
    if len(gst) != 15:
        logger.warning(f"Invalid GST length: {gst} (length: {len(gst)})")
        return ""
    
    # Check pattern
    if not GST_PATTERN.match(gst):
        logger.warning(f"Invalid GST pattern: {gst}")
        return ""
    
    # Check state code
    state_code = gst[:2]
    if state_code not in VALID_STATE_CODES:
        logger.warning(f"Invalid GST state code: {state_code} in {gst}")
        return ""
    
    return gst


def validate_phone_number(phone: str) -> str:
    """Validate and clean phone number."""
    if not phone:
        return ""
    
    # Remove common separators
    cleaned = re.sub(r'[\s\-\(\)]', '', phone)
    
    # Check if it matches Indian phone pattern
    if PHONE_PATTERN.search(phone):
        return phone.strip()
    
    # Check if it's 10 digits
    if len(cleaned) == 10 and cleaned.isdigit() and cleaned[0] in '6789':
        return phone.strip()
    
    # Check if it's 12 digits with country code
    if len(cleaned) == 12 and cleaned.startswith('91') and cleaned[2] in '6789':
        return phone.strip()
    
    logger.warning(f"Invalid phone number format: {phone}")
    return ""


def validate_imei(imei: str) -> str:
    """Validate IMEI number (15 digits)."""
    if not imei:
        return ""
    
    cleaned = re.sub(r'[^\d]', '', imei)
    
    if len(cleaned) == 15 and cleaned.isdigit():
        return cleaned
    
    logger.warning(f"Invalid IMEI: {imei}")
    return ""


def validate_date(date_str: str) -> str:
    """Validate date format."""
    if not date_str:
        return ""
    
    # Check if matches common date patterns
    for pattern in DATE_PATTERNS:
        if re.search(pattern, date_str):
            return date_str.strip()
    
    logger.warning(f"Invalid date format: {date_str}")
    return ""


def validate_tax_percentage(value) -> str:
    """Validate tax percentage against valid GST rates."""
    if not value or value == "":
        return ""
    
    try:
        # Extract numeric value
        num_value = float(re.sub(r"[^\d.]+", "", str(value)))
        
        # Check if in valid list (with tolerance)
        for valid_pct in VALID_TAX_PERCENTAGES:
            if abs(num_value - valid_pct) < 0.01:
                return num_value
        
        logger.warning(f"Invalid tax percentage: {value} (not in valid list)")
        return ""
    except:
        return ""


def validate_invoice_number(inv_num: str) -> str:
    """Validate invoice number."""
    if not inv_num:
        return ""
    
    cleaned = inv_num.strip()
    
    # Must have at least one alphanumeric character
    if not re.search(r'[A-Za-z0-9]', cleaned):
        logger.warning(f"Invalid invoice number: {inv_num}")
        return ""
    
    return cleaned


##########################################
### CROSS-FIELD VALIDATION
##########################################

def cross_validate_extraction(data: Dict) -> Dict:
    """Perform cross-field validation and corrections."""
    issues = []
    
    # Validate GST number
    if data.get("gstNumber"):
        validated_gst = validate_gst_number(data["gstNumber"])
        if validated_gst != data["gstNumber"]:
            issues.append(f"GST corrected: {data['gstNumber']} → {validated_gst}")
            data["gstNumber"] = validated_gst
    
    # Validate phone number
    if data.get("customerPhone"):
        validated_phone = validate_phone_number(data["customerPhone"])
        if validated_phone != data["customerPhone"]:
            issues.append(f"Phone corrected: {data['customerPhone']} → {validated_phone}")
            data["customerPhone"] = validated_phone
    
    # Validate invoice number
    if data.get("invoiceNumber"):
        validated_inv = validate_invoice_number(data["invoiceNumber"])
        if validated_inv != data["invoiceNumber"]:
            issues.append(f"Invoice# corrected: {data['invoiceNumber']} → {validated_inv}")
            data["invoiceNumber"] = validated_inv
    
    # Validate date
    if data.get("invoiceDate"):
        validated_date = validate_date(data["invoiceDate"])
        if validated_date != data["invoiceDate"]:
            issues.append(f"Date corrected: {data['invoiceDate']} → {validated_date}")
            data["invoiceDate"] = validated_date
    
    # Validate items
    for idx, item in enumerate(data.get("items", [])):
        # Validate IMEI
        if item.get("imeiNumber"):
            validated_imei = validate_imei(item["imeiNumber"])
            if validated_imei != item["imeiNumber"]:
                issues.append(f"Item {idx+1} IMEI corrected")
                item["imeiNumber"] = validated_imei
        
        # Validate tax percentages
        for tax_field in ["sgst", "cgst", "igst"]:
            if item.get(tax_field):
                validated_tax = validate_tax_percentage(item[tax_field])
                if str(validated_tax) != str(item[tax_field]):
                    issues.append(f"Item {idx+1} {tax_field} corrected: {item[tax_field]} → {validated_tax}")
                    item[tax_field] = validated_tax
        
        # Validate tax mutual exclusivity
        has_sgst_cgst = item.get("sgst") and item.get("cgst")
        has_igst = item.get("igst")
        
        if has_sgst_cgst and has_igst:
            issues.append(f"Item {idx+1}: Both SGST/CGST and IGST present - clearing IGST")
            item["igst"] = ""
    
    if issues:
        logger.info(f"Validation issues fixed: {', '.join(issues)}")
    
    return data


##########################################
### CONFIDENCE SCORING
##########################################

def calculate_field_confidence(data: Dict) -> Dict:
    """Calculate confidence score for each field."""
    confidence = {}
    
    # Critical fields (must have)
    critical_fields = ["invoiceNumber", "invoiceDate", "supplierName"]
    for field in critical_fields:
        confidence[field] = 100 if data.get(field) else 0
    
    # GST number (validate format)
    if data.get("gstNumber"):
        gst = data["gstNumber"]
        if len(gst) == 15 and GST_PATTERN.match(gst):
            confidence["gstNumber"] = 100
        else:
            confidence["gstNumber"] = 50
    else:
        confidence["gstNumber"] = 0
    
    # Items completeness
    items = data.get("items", [])
    if items:
        item_scores = []
        for item in items:
            required = ["description", "quantity", "rate"]
            score = sum(100 for f in required if item.get(f)) / len(required)
            item_scores.append(score)
        confidence["items"] = sum(item_scores) / len(item_scores)
    else:
        confidence["items"] = 0
    
    # Overall confidence
    confidence["overall"] = sum(confidence.values()) / len(confidence)
    
    return confidence


##########################################
### NEW: ITEM-LEVEL CONFIDENCE SCORING
##########################################

def calculate_item_amount_confidence(item: Dict, calculated_amount: float) -> int:
    """Calculate confidence score for extracted itemAmount."""
    extracted_amount = parse_amount(item.get("itemAmount"))
    rate = parse_amount(item.get("rate"))
    quantity = parse_amount(item.get("quantity")) or 1
    
    confidence = 100
    
    # Check 1: Is extracted amount reasonable compared to calculated?
    if extracted_amount > 0 and calculated_amount > 0:
        diff_ratio = abs(extracted_amount - calculated_amount) / calculated_amount
        if diff_ratio > 0.15:  # More than 15% difference
            confidence -= 30
            logger.debug(f"Item amount differs from calculated by {diff_ratio*100:.1f}%")
    
    # Check 2: Is amount > rate? (should be for most cases with tax)
    if extracted_amount > 0 and rate > 0 and quantity >= 1:
        if extracted_amount < rate and quantity > 1:
            confidence -= 40
            logger.debug(f"Item amount ({extracted_amount}) less than rate ({rate}) for qty {quantity}")
    
    # Check 3: Does it have proper decimal places? (amounts usually do)
    if extracted_amount > 0:
        decimal_part = extracted_amount % 1
        if decimal_part == 0 and rate % 1 != 0:
            confidence -= 20
            logger.debug(f"Item amount has no decimals but rate does")
    
    # Check 4: Is extracted amount zero but we have rate and quantity?
    if extracted_amount == 0 and rate > 0 and quantity > 0:
        confidence = 0
        logger.debug(f"Item amount is zero but rate and quantity exist")
    
    return max(0, confidence)


##########################################
### ENHANCED EXTRACTION FUNCTION
##########################################

def extract_invoice(image_path: str, prompt: str, retries: int = 3) -> Dict:
    """Extract invoice data with enhanced error handling and validation."""
    
    logger.info(f"Processing: {os.path.basename(image_path)}")
    
    with open(image_path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode("utf-8")
    
    mime = "image/jpeg" if image_path.lower().endswith((".jpg", ".jpeg")) else "image/png"
    data_url = f"data:{mime};base64,{img_b64}"
    
    messages = [{
        "role": "user",
        "content": [
            {"type": "image_url", "image_url": {"url": data_url}},
            {"type": "text", "text": prompt.strip()}
        ]
    }]
    
    # Retry with exponential backoff
    for attempt in range(retries + 1):
        try:
            response = model.chat(messages=messages, params=generation_params)
            raw = response["choices"][0]["message"]["content"].strip()
            break
        except Exception as e:
            if attempt < retries:
                wait_time = (2 ** attempt) + random.uniform(0, 1)
                logger.warning(f"Retry {attempt+1}/{retries} after {wait_time:.1f}s: {e}")
                time.sleep(wait_time)
            else:
                logger.error(f"Failed after {retries} retries: {e}")
                raise
    
    # Parse JSON with multiple strategies
    data = parse_json_robust(raw, image_path)
    
    # Validate and correct
    data = cross_validate_extraction(data)
    
    # Calculate confidence
    confidence = calculate_field_confidence(data)
    data["_confidence"] = confidence
    
    # Log confidence
    logger.info(f"Overall confidence: {confidence['overall']:.1f}%")
    
    data["Image_File"] = os.path.basename(image_path)
    return data


def parse_json_robust(raw_text: str, image_path: str) -> Dict:
    """Parse JSON with multiple fallback strategies."""
    
    # Strategy 1: Direct parse
    try:
        return json.loads(raw_text)
    except:
        pass
    
    # Strategy 2: Remove markdown
    cleaned = re.sub(r'^```(?:json)?\s*|\s*```$', '', raw_text, flags=re.DOTALL).strip()
    
    # Strategy 3: Extract JSON block
    start, end = cleaned.find("{"), cleaned.rfind("}")
    
    if start != -1 and end > start:
        json_str = cleaned[start:end+1]
        
        # Strategy 4: Fix common issues
        fixes = [
            (r',\s*}', '}'),  # Trailing commas in objects
            (r',\s*]', ']'),  # Trailing commas in arrays
            (r"'", '"'),  # Single quotes to double
            (r'\bTrue\b', 'true'),
            (r'\bFalse\b', 'false'),
            (r'\bNone\b', 'null'),
        ]
        
        for pattern, replacement in fixes:
            json_str = re.sub(pattern, replacement, json_str)
        
        try:
            return json.loads(json_str)
        except Exception as e:
            logger.error(f"JSON parse failed for {os.path.basename(image_path)}: {e}")
    
    # Return empty structure if all fails
    logger.error(f"Could not parse JSON from {os.path.basename(image_path)}")
    return {"items": []}


##########################################
### UTILITY FUNCTIONS
##########################################

def parse_amount(value) -> float:
    """Parse amount from string."""
    try:
        return float(re.sub(r"[^\d.]+", "", str(value))) if value else 0.0
    except:
        return 0.0


def detect_tax_inclusive_invoice(items: List[Dict]) -> bool:
    """Detect if invoice uses tax-inclusive rates."""
    matches = 0
    total_comparisons = 0
    
    for item in items:
        extracted_rate = parse_amount(item.get("rate"))
        extracted_amount = parse_amount(item.get("itemAmount"))
        
        if extracted_rate > 0 and extracted_amount > 0:
            total_comparisons += 1
            difference_ratio = abs(extracted_amount - extracted_rate) / extracted_rate
            
            if difference_ratio < 0.01:
                matches += 1
    
    if total_comparisons > 0 and (matches / total_comparisons) >= 0.5:
        return True
    
    return False


##########################################
### NEW: ENHANCED MULTI-LINE CALCULATION
##########################################

def calculate_item_amount_enhanced(item: Dict, is_tax_inclusive: bool = False, 
                                   net_total: float = None, is_single_item: bool = False) -> float:
    """
    Enhanced calculation for item amount with two-pass strategy for multi-line invoices.
    
    Strategy:
    Pass 1: Use extracted value if confidence is high
    Pass 2: Calculate from rate if Pass 1 fails
    Pass 3: Will be applied later for proportional adjustment
    """
    extracted_rate = parse_amount(item.get("rate"))
    extracted_amount = parse_amount(item.get("itemAmount"))
    quantity = parse_amount(item.get("quantity")) or 1
    
    # For single item invoices, use net total
    if is_single_item and net_total and net_total > 0:
        logger.debug(f"Single item: Using net total {net_total}")
        return round(net_total, 2)
    
    # Get validated tax percentages
    def to_percent(val):
        validated = validate_tax_percentage(val)
        return float(validated) if validated != "" else 0.0
    
    sgst_p = to_percent(item.get("sgst", ""))
    cgst_p = to_percent(item.get("cgst", ""))
    igst_p = to_percent(item.get("igst", ""))
    combined_tax = sgst_p + cgst_p if sgst_p or cgst_p else igst_p
    
    # Calculate expected amount from rate
    if extracted_rate > 0:
        if is_tax_inclusive:
            calculated_amount = extracted_rate * quantity
        else:
            calculated_amount = extracted_rate * quantity * (1 + combined_tax / 100)
    else:
        calculated_amount = 0.0
    
    # PASS 1: Check if extracted amount is reliable
    if extracted_amount > 0:
        confidence = calculate_item_amount_confidence(item, calculated_amount)
        
        if confidence >= 70:
            # High confidence in extracted amount
            logger.debug(f"Pass 1: Using extracted amount {extracted_amount} (confidence: {confidence}%)")
            item["_amount_source"] = "extracted"
            item["_amount_confidence"] = confidence
            return round(extracted_amount, 2)
        else:
            logger.debug(f"Pass 1: Low confidence {confidence}% in extracted amount {extracted_amount}")
    
    # PASS 2: Calculate from rate
    if extracted_rate > 0:
        logger.debug(f"Pass 2: Calculating from rate {extracted_rate} × {quantity} × (1 + {combined_tax}%)")
        item["_amount_source"] = "calculated"
        item["_amount_confidence"] = 85
        
        # Additional validation: Check if rate makes sense
        if extracted_amount > 0:
            # We have both rate and amount, check consistency
            if quantity > 1 and extracted_amount < extracted_rate:
                # Amount shouldn't be less than rate for multiple quantities
                logger.warning(f"Inconsistent: amount {extracted_amount} < rate {extracted_rate} for qty {quantity}")
                # Use calculated value
                return round(calculated_amount, 2)
        
        return round(calculated_amount, 2)
    
    # PASS 3: Fallback to extracted amount if available
    if extracted_amount > 0:
        logger.debug(f"Pass 3: Using extracted amount {extracted_amount} as fallback")
        item["_amount_source"] = "extracted_fallback"
        item["_amount_confidence"] = 50
        return round(extracted_amount, 2)
    
    # No valid data
    logger.warning(f"No valid amount data for item: {item.get('description', 'Unknown')}")
    item["_amount_source"] = "none"
    item["_amount_confidence"] = 0
    return 0.0


def apply_proportional_adjustment(items: List[Dict], net_total: float) -> List[Dict]:
    """
    Apply proportional adjustment to item amounts to match net total.
    This is Pass 3 of the calculation strategy.
    """
    if not items or net_total <= 0:
        return items
    
    # Calculate sum of current item amounts
    current_total = sum(parse_amount(item.get("itemAmount") or 0) for item in items)
    
    if current_total <= 0:
        logger.warning("Cannot apply proportional adjustment: current total is 0")
        return items
    
    # Calculate adjustment ratio
    adjustment_ratio = net_total / current_total
    difference_pct = abs(1 - adjustment_ratio) * 100
    
    # Only adjust if difference is significant but not too large
    if difference_pct > 5 and difference_pct < 20:
        logger.info(f"Applying proportional adjustment: {difference_pct:.2f}% difference")
        
        for item in items:
            original_amount = parse_amount(item.get("itemAmount") or 0)
            adjusted_amount = original_amount * adjustment_ratio
            item["itemAmount"] = round(adjusted_amount, 2)
            item["_adjusted"] = True
            logger.debug(f"Adjusted: {original_amount} → {adjusted_amount}")
    elif difference_pct >= 20:
        logger.warning(f"Difference too large ({difference_pct:.2f}%) - skipping adjustment")
    else:
        logger.debug(f"Difference acceptable ({difference_pct:.2f}%) - no adjustment needed")
    
    return items


def validate_multi_line_totals(items: List[Dict], net_total: float) -> Dict:
    """
    Validate that sum of item amounts matches net total.
    Returns validation results and statistics.
    """
    if not items:
        return {"valid": False, "reason": "no_items"}
    
    calculated_total = sum(parse_amount(item.get("itemAmount") or 0) for item in items)
    
    if net_total <= 0:
        return {
            "valid": True,
            "reason": "no_net_total",
            "calculated_total": calculated_total,
            "difference": 0
        }
    
    difference = abs(calculated_total - net_total)
    difference_pct = (difference / net_total) * 100
    
    is_valid = difference_pct <= 5  # Allow 5% tolerance
    
    return {
        "valid": is_valid,
        "calculated_total": round(calculated_total, 2),
        "net_total": round(net_total, 2),
        "difference": round(difference, 2),
        "difference_pct": round(difference_pct, 2),
        "reason": "within_tolerance" if is_valid else "exceeds_tolerance"
    }


def calculate_item_amount(item: Dict, is_tax_inclusive: bool = False, 
                         net_total: float = None, is_single_item: bool = False) -> float:
    """Calculate final item amount with tax logic."""
    extracted_rate = parse_amount(item.get("rate"))
    extracted_amount = parse_amount(item.get("itemAmount"))
    quantity = parse_amount(item.get("quantity")) or 1
    
    # For single item invoices, use net total
    if is_single_item and net_total and net_total > 0:
        return round(net_total, 2)
    
    # Get validated tax percentages
    def to_percent(val):
        validated = validate_tax_percentage(val)
        return float(validated) if validated != "" else 0.0
    
    sgst_p = to_percent(item.get("sgst", ""))
    cgst_p = to_percent(item.get("cgst", ""))
    igst_p = to_percent(item.get("igst", ""))
    combined_tax = sgst_p + cgst_p if sgst_p or cgst_p else igst_p
    
    # If rate and amount match, it's tax-inclusive
    if extracted_rate > 0 and extracted_amount > 0:
        difference_ratio = abs(extracted_amount - extracted_rate) / extracted_rate
        if difference_ratio < 0.01:
            return round(extracted_rate * quantity, 2)
    
    # Tax-inclusive invoice
    if is_tax_inclusive and extracted_rate > 0:
        return round(extracted_rate * quantity, 2)
    
    # Tax-exclusive calculation
    if extracted_rate > 0 and combined_tax > 0:
        final_amount = extracted_rate * quantity * (1 + combined_tax / 100)
    elif extracted_rate > 0:
        final_amount = extracted_rate * quantity
    else:
        final_amount = extracted_amount if extracted_amount > 0 else 0.0
    
    return round(final_amount, 2)


def calculate_string_similarity(str1: str, str2: str) -> float:
    """Calculate Levenshtein similarity score (0-100)."""
    if not str1 or not str2:
        return 0
    
    s1 = str(str1).lower().strip()
    s2 = str(str2).lower().strip()
    
    if s1 == s2:
        return 100
    
    def levenshtein_distance(s1, s2):
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    distance = levenshtein_distance(s1, s2)
    max_len = max(len(s1), len(s2))
    similarity = (1 - distance / max_len) * 100
    
    return round(similarity, 2)


##########################################
### MAIN PROCESS WITH ENHANCED LOGIC
##########################################

def main():
    """Main extraction process with enhanced multi-line validation."""
    
    files = [os.path.join(IMAGES_DIR, f) for f in os.listdir(IMAGES_DIR) 
             if os.path.splitext(f)[1].lower() in (".jpg", ".jpeg", ".png")]
    
    if not files:
        logger.error("No invoice images found!")
        return
    
    logger.info(f"Found {len(files)} images to process")
    
    results = []
    failed = []
    
    # Process with thread pool
    with ThreadPoolExecutor(max_workers=min(4, len(files))) as executor:
        futures = {executor.submit(extract_invoice, path, invoice_prompt, retries=3): path 
                  for path in files}
        
        for n, future in enumerate(as_completed(futures), start=1):
            path = futures[future]
            try:
                data = future.result()
                confidence = data.get("_confidence", {}).get("overall", 0)
                status = "✓" if confidence > 70 else "⚠"
                print(f"[{status}{n}/{len(files)}] {os.path.basename(path)} - Confidence: {confidence:.1f}%")
                results.append(data)
            except Exception as e:
                logger.error(f"Failed: {os.path.basename(path)} - {e}")
                failed.append(os.path.basename(path))
    
    if failed:
        logger.warning(f"Failed to process {len(failed)} files: {', '.join(failed)}")
    
    # Process results
    expanded_rows, summary_rows = [], []
    
    for record in results:
        items = record.get("items", [])
        image_file = record.get("Image_File")
        confidence = record.get("_confidence", {})
        
        # Extract header fields
        invoice_number = record.get("invoiceNumber", "")
        invoice_number_type = record.get("invoiceNumberType", "")
        invoice_date = record.get("invoiceDate", "")
        supplier_name = record.get("supplierName", "")
        gst_number = record.get("gstNumber", "")
        customer_name = record.get("customerName", "")
        customer_phone = record.get("customerPhone", "")
        customer_address = record.get("customerAddress", "")
        down_payment = record.get("downPayment", "")
        net_total_raw = parse_amount(record.get("netTotal") or 0)
        
        # Stamp information
        stamp_present = record.get("stampPresent", "")
        information_in_stamp = record.get("informationInStamp", "")
        signature_present = record.get("signaturePresent", "")
        hypothecation_stamp = record.get("hypothecationStamp", "")
        stamp_matching_score = record.get("stampCompanyMatching_score", 0)
        
        # Calculate stamp matching if not provided
        if stamp_present == "Present" and information_in_stamp and stamp_matching_score == 0:
            stamp_matching_score = calculate_string_similarity(supplier_name, information_in_stamp)
        
        # Detect invoice type
        is_single_item = len(items) == 1
        is_tax_inclusive = detect_tax_inclusive_invoice(items)
        
        if is_tax_inclusive:
            logger.info(f"Tax-inclusive invoice detected: {image_file}")
        
        # NEW: Use enhanced calculation for multi-line invoices
        if len(items) > 1:
            logger.info(f"Multi-line invoice detected: {len(items)} items")
        
        # Process items with enhanced calculation
        for item in items:
            # Clean description
            desc = re.sub(r"\bTVS\b", "", str(item.get("description", "")), flags=re.IGNORECASE)
            desc = re.sub(r"IMEI\s*\d*\s*[:\-]?\s*\d{5,20}", "", desc, flags=re.IGNORECASE)
            desc = re.sub(r"IMEI\s*\d*\s*[:\-]?", "", desc, flags=re.IGNORECASE)
            desc = re.sub(r"\s{2,}", " ", desc).strip()
            item["description"] = desc
            
            # Validate tax percentages
            item["sgst"] = validate_tax_percentage(item.get("sgst", ""))
            item["cgst"] = validate_tax_percentage(item.get("cgst", ""))
            item["igst"] = validate_tax_percentage(item.get("igst", ""))
            
            # Calculate combined tax
            sgst_p = float(item["sgst"]) if item["sgst"] != "" else 0.0
            cgst_p = float(item["cgst"]) if item["cgst"] != "" else 0.0
            igst_p = float(item["igst"]) if item["igst"] != "" else 0.0
            combined_tax = sgst_p + cgst_p if sgst_p or cgst_p else igst_p
            item["tax"] = combined_tax
            
            # NEW: Use enhanced calculation for multi-line
            if len(items) > 1:
                final_amount = calculate_item_amount_enhanced(item, is_tax_inclusive, net_total_raw, is_single_item)
            else:
                final_amount = calculate_item_amount(item, is_tax_inclusive, net_total_raw, is_single_item)
            
            item["itemAmount"] = final_amount
            
            # Add to expanded rows
            expanded_rows.append({
                "Image_File": image_file,
                "invoiceNumber": invoice_number,
                "invoiceNumberType": invoice_number_type,
                "invoiceDate": invoice_date,
                "supplierName": supplier_name,
                "gstNumber": gst_number,
                "customerName": customer_name,
                "customerPhone": customer_phone,
                "customerAddress": customer_address,
                "downPayment": down_payment,
                "stampPresent": stamp_present,
                "informationInStamp": information_in_stamp,
                "signaturePresent": signature_present,
                "hypothecationStamp": hypothecation_stamp,
                "stampCompanyMatching_score": stamp_matching_score,
                "itemNo": item.get("itemNo", ""),
                "description": item.get("description", ""),
                "brandName": item.get("brandName", ""),
                "imeiNumber": item.get("imeiNumber", ""),
                "serialNumber": item.get("serialNumber", ""),
                "quantity": item.get("quantity", ""),
                "rate": item.get("rate", ""),
                "sgst": item.get("sgst", ""),
                "cgst": item.get("cgst", ""),
                "igst": item.get("igst", ""),
                "tax": item.get("tax", ""),
                "itemAmount": item.get("itemAmount", "")
            })
        
        # NEW: Validate multi-line totals
        if len(items) > 1 and net_total_raw > 0:
            validation = validate_multi_line_totals(items, net_total_raw)
            logger.info(f"Multi-line validation for {image_file}: {validation}")
            
            # Apply proportional adjustment if needed
            if not validation["valid"] and validation["difference_pct"] > 5:
                logger.info(f"Applying proportional adjustment for {image_file}")
                items = apply_proportional_adjustment(items, net_total_raw)
                
                # Update expanded rows with adjusted amounts
                item_idx = len(expanded_rows) - len(items)
                for i, item in enumerate(items):
                    expanded_rows[item_idx + i]["itemAmount"] = item.get("itemAmount", "")
        
        # Calculate total
        if is_single_item:
            net_total = net_total_raw if net_total_raw > 0 else parse_amount(items[0].get("itemAmount") or 0)
        else:
            net_total = sum(parse_amount(item.get("itemAmount") or 0) for item in items)
        
        # Add to summary
        summary_rows.append({
            "Image_File": image_file,
            "invoiceNumber": invoice_number,
            "invoiceNumberType": invoice_number_type,
            "invoiceDate": invoice_date,
            "supplierName": supplier_name,
            "gstNumber": gst_number,
            "customerName": customer_name,
            "customerPhone": customer_phone,
            "customerAddress": customer_address,
            "downPayment": down_payment,
            "stampPresent": stamp_present,
            "informationInStamp": information_in_stamp,
            "signaturePresent": signature_present,
            "hypothecationStamp": hypothecation_stamp,
            "stampCompanyMatching_score": stamp_matching_score,
            "Total_Amount": round(net_total, 2),
            "Tax_Inclusive": "Yes" if is_tax_inclusive else "No"
        })
    
    # Export to Excel
    if expanded_rows:
        df_items = pd.DataFrame(expanded_rows)
        df_summary = pd.DataFrame(summary_rows)
        os.makedirs(os.path.dirname(EXCEL_OUTPUT_PATH), exist_ok=True)
        
        with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
            df_items.to_excel(writer, index=False, sheet_name="Item_Details")
            df_summary.to_excel(writer, index=False, sheet_name="Invoice_Summary")
            
            # Format as tables
            wb = writer.book
            ws_items = writer.sheets["Item_Details"]
            ws_summary = writer.sheets["Invoice_Summary"]
            
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter
            
            items_range = f"A1:{get_column_letter(df_items.shape[1])}{df_items.shape[0] + 1}"
            summary_range = f"A1:{get_column_letter(df_summary.shape[1])}{df_summary.shape[0] + 1}"
            
            table_items = Table(displayName="ItemDetailsTable", ref=items_range)
            table_summary = Table(displayName="InvoiceSummaryTable", ref=summary_range)
            
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )
            table_items.tableStyleInfo = style
            table_summary.tableStyleInfo = style
            
            ws_items.add_table(table_items)
            ws_summary.add_table(table_summary)
        
        logger.info(f"✅ Saved results to {EXCEL_OUTPUT_PATH}")
        logger.info(f"📊 Processed {len(results)} invoices with {len(expanded_rows)} total items")
        
        # Print quality summary
        avg_confidence = sum(r.get("_confidence", {}).get("overall", 0) for r in results) / len(results)
        logger.info(f"📈 Average confidence: {avg_confidence:.1f}%")
    else:
        logger.error("No invoice data extracted!")


if __name__ == "__main__":
    main()